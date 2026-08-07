# services/export.py
import pandas as pd
from database.db import Database
from datetime import datetime
import io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

class ExportService:
    @staticmethod
    def export_to_excel(user_id):
        """Экспорт всех сделок пользователя в Excel с форматированием"""
        trades = Database.get_all_trades(user_id)
        
        if not trades:
            return None
        
        # Подготовка данных
        data = []
        for trade in trades:
            # Определяем цвет для результата
            result_color = ""
            if trade.status == 'closed':
                if trade.result == 'profit':
                    result_color = "🟢"
                elif trade.result == 'loss':
                    result_color = "🔴"
                elif trade.result == 'breakeven':
                    result_color = "⚪"
            else:
                result_color = "🟡"
            
            # Форматируем PnL
            pnl_display = ""
            if trade.pnl is not None:
                pnl_display = f"{'+' if trade.pnl > 0 else ''}{trade.pnl:.2f}"
            
            data.append({
                'ID': trade.id,
                'Дата открытия': trade.date.strftime('%d.%m.%Y %H:%M'),
                'Монета': trade.symbol,
                'Направление': trade.direction,
                'Размер позиции': f"{trade.position_size:.2f}$",
                'Цена входа': f"{trade.entry_price:.4f}" if trade.entry_price else '',
                'Цена выхода': f"{trade.exit_price:.4f}" if trade.exit_price else '',
                'Сетап': trade.setup or '',
                'Уверенность': f"{trade.confidence}/10" if trade.confidence else '',
                'Статус': 'Открыта 🟡' if trade.status == 'open' else 'Закрыта',
                'Результат': result_color,
                'PnL': pnl_display,
                'Депозит': f"{trade.deposit:.2f}$" if trade.deposit else '',
                'Дата закрытия': trade.close_date.strftime('%d.%m.%Y %H:%M') if trade.close_date else '',
                'Ошибка': trade.mistake or ''
            })
        
        df = pd.DataFrame(data)
        
        # Создаем Excel файл с форматированием
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Основной лист со сделками
            df.to_excel(writer, sheet_name='Сделки', index=False)
            
            # Получаем workbook для форматирования
            workbook = writer.book
            worksheet = writer.sheets['Сделки']
            
            # Настройка стилей
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            # Применяем стили к заголовкам
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Настройка ширины колонок
            column_widths = {
                'A': 8,   # ID
                'B': 18,  # Дата открытия
                'C': 12,  # Монета
                'D': 12,  # Направление
                'E': 16,  # Размер позиции
                'F': 14,  # Цена входа
                'G': 14,  # Цена выхода
                'H': 40,  # Сетап
                'I': 14,  # Уверенность
                'J': 14,  # Статус
                'K': 12,  # Результат
                'L': 14,  # PnL
                'M': 14,  # Депозит
                'N': 18,  # Дата закрытия
                'O': 30,  # Ошибка
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # Цветовая подсветка для PnL
            for row in range(2, len(df) + 2):
                pnl_cell = worksheet[f'L{row}']
                if pnl_cell.value:
                    try:
                        value = float(pnl_cell.value)
                        if value > 0:
                            pnl_cell.font = Font(color='008000', bold=True)
                        elif value < 0:
                            pnl_cell.font = Font(color='FF0000', bold=True)
                    except:
                        pass
            
            # Добавляем лист со статистикой
            stats_data = ExportService._get_stats_data(user_id)
            if stats_data:
                stats_df = pd.DataFrame(stats_data)
                stats_df.to_excel(writer, sheet_name='Статистика', index=False)
                
                # Форматируем статистику
                stats_worksheet = writer.sheets['Статистика']
                for cell in stats_worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                stats_worksheet.column_dimensions['A'].width = 25
                stats_worksheet.column_dimensions['B'].width = 20
            
            # Добавляем лист с анализом по сетапам
            setups_data = ExportService._get_setups_data(user_id)
            if setups_data:
                setups_df = pd.DataFrame(setups_data)
                setups_df.to_excel(writer, sheet_name='Сетапы', index=False)
                
                setups_worksheet = writer.sheets['Сетапы']
                for cell in setups_worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                for col in ['A', 'B', 'C', 'D', 'E']:
                    setups_worksheet.column_dimensions[col].width = 20
        
        output.seek(0)
        return output
    
    @staticmethod
    def _get_stats_data(user_id):
        """Получение данных для листа статистики"""
        stats = Database.get_stats(user_id)
        user = Database.get_or_create_user(user_id)
        
        return [
            ['Показатель', 'Значение'],
            ['Всего сделок', stats['total']],
            ['Побед', stats['wins']],
            ['Поражений', stats['losses']],
            ['Безубыточных', stats['breakevens']],
            ['Win Rate', f"{stats['win_rate']:.2f}%"],
            ['Общий PnL', f"{stats['total_pnl']:.2f}$"],
            ['Начальный депозит', f"{user.initial_deposit:.2f}$"],
            ['Текущий депозит', f"{stats['current_deposit']:.2f}$"],
            ['Средняя прибыль', f"+{stats['avg_profit']:.2f}$"],
            ['Средний убыток', f"-{abs(stats['avg_loss']):.2f}$"],
            ['Profit Factor', f"{stats['profit_factor']:.2f}"],
            ['Макс. серия побед', stats['max_wins']],
            ['Макс. серия убытков', stats['max_losses']]
        ]
    
    @staticmethod
    def _get_setups_data(user_id):
        """Получение данных для листа сетапов"""
        setups = Database.get_stats_by_setup(user_id)
        
        if not setups:
            return None
        
        data = [['Сетап', 'Всего', 'Побед', 'Win Rate', 'PnL']]
        for setup, stats in setups.items():
            data.append([
                setup,
                stats['total'],
                stats['wins'],
                f"{stats['win_rate']:.1f}%",
                f"{stats['pnl']:.2f}$"
            ])
        
        return data

    @staticmethod
    def export_trade_template():
        """Экспорт шаблона для заполнения сделок в Excel"""
        output = io.BytesIO()
        
        wb = Workbook()
        
        # Основной лист
        ws = wb.active
        ws.title = "Шаблон сделок"
        
        # Заголовки
        headers = [
            'Дата', 'Монета', 'Направление', 'Размер позиции',
            'Цена входа', 'Цена выхода', 'Сетап', 'Уверенность',
            'Результат', 'PnL', 'Ошибка'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Примеры данных
        examples = [
            ['07.08.2026 14:30', 'BTCUSDT', 'LONG', 500, 58000, 59000, 'Ложный пробой уровня', 8, 'profit', '+100', ''],
            ['07.08.2026 15:45', 'ETHUSDT', 'SHORT', 300, 3200, 3150, 'Отбой от VWAP', 7, 'loss', '-50', 'Ранний вход'],
        ]
        
        for row, example in enumerate(examples, 2):
            for col, value in enumerate(example, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Настройка ширины колонок
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 18
        
        # Лист с описанием
        ws_info = wb.create_sheet("Инструкция")
        info_data = [
            ['📖 Инструкция по заполнению шаблона', ''],
            ['', ''],
            ['Поле', 'Описание'],
            ['Дата', 'Формат: ДД.ММ.ГГГГ ЧЧ:ММ (например: 07.08.2026 14:30)'],
            ['Монета', 'Тикер монеты (например: BTCUSDT, ETHUSDT)'],
            ['Направление', 'LONG или SHORT'],
            ['Размер позиции', 'В долларах (например: 500)'],
            ['Цена входа', 'Цена входа в сделку'],
            ['Цена выхода', 'Цена выхода из сделки'],
            ['Сетап', 'Ваша стратегия или описание'],
            ['Уверенность', 'Оценка от 1 до 10'],
            ['Результат', 'profit, loss или breakeven'],
            ['PnL', 'Прибыль/убыток в долларах (например: +100 или -50)'],
            ['Ошибка', 'Причина ошибки (опционально)'],
        ]
        
        for row, data in enumerate(info_data, 1):
            for col, value in enumerate(data, 1):
                ws_info.cell(row=row, column=col, value=value)
        
        for col in range(1, 3):
            ws_info.column_dimensions[chr(64 + col)].width = 30
        
        wb.save(output)
        output.seek(0)
        return output